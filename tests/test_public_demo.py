import runpy
import sys
import tempfile
import types
import unittest
import json
from pathlib import Path
from unittest.mock import patch

import openpyxl

from workbench_config import ConfigurationError, load_web_urls


PROJECT_DIR = Path(__file__).resolve().parents[1]
GUI_SCRIPT = PROJECT_DIR / "procurement_workbench.py"


class PublicDemoConfigurationTests(unittest.TestCase):
    def setUp(self):
        ttk_stub = types.ModuleType("ttkbootstrap")
        with patch.dict(sys.modules, {"ttkbootstrap": ttk_stub}):
            self.module = runpy.run_path(str(GUI_SCRIPT))

    def test_categories_use_obvious_synthetic_identifiers(self):
        categories = self.module["CATEGORIES"]

        self.assertIn("Demo Category A", categories)
        self.assertGreaterEqual(len(categories), 3)
        for category_name, config in categories.items():
            self.assertTrue(category_name.startswith("Demo Category"))
            self.assertTrue(config["label"].startswith("Demo Category"))
            self.assertTrue(config["data_sheet"].startswith("DEMO_"))
            for material_id in config["filter_materials"]:
                self.assertTrue(str(material_id).startswith("DEMO-MAT-"))

    def test_demo_material_can_be_filtered_without_network_access(self):
        category = "Demo Category B"
        config = self.module["CATEGORIES"][category]

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "demo_purchase_records.xlsx"
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Sheet1"
            worksheet.append(["物料", "订单净值", "短文本", "净价", "采购订单数量"])
            worksheet.append([config["filter_materials"][0], 240, "DEMO-MATERIAL", 160, 2])
            worksheet.append(["DEMO-MAT-UNRELATED", 300, "UNRELATED", 120, 1])
            workbook.save(workbook_path)
            workbook.close()

            function_globals = self.module["_enhance_and_filter"].__globals__
            with patch.dict(function_globals, {"EXCEL_FILE": str(workbook_path)}):
                _, row_count, _ = self.module["_enhance_and_filter"](
                    config, lambda _message: None
                )

            self.assertEqual(row_count, 1)
            workbook = openpyxl.load_workbook(workbook_path, data_only=True)
            self.assertIn(config["data_sheet"], workbook.sheetnames)
            self.assertEqual(workbook[config["data_sheet"]].cell(2, 1).value, config["filter_materials"][0])
            workbook.close()

    def test_generated_data_sheets_follow_the_synthetic_category_order(self):
        category_b = self.module["CATEGORIES"]["Demo Category B"]
        category_c = self.module["CATEGORIES"]["Demo Category C"]

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "demo_purchase_records.xlsx"
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Sheet1"
            worksheet.append(["物料", "订单净值", "短文本", "净价", "采购订单数量"])
            worksheet.append([category_b["filter_materials"][0], 240, "DEMO-B", 160, 2])
            worksheet.append([category_c["filter_materials"][0], 360, "DEMO-C", 120, 3])
            workbook.save(workbook_path)
            workbook.close()

            function_globals = self.module["_enhance_and_filter"].__globals__
            with patch.dict(function_globals, {"EXCEL_FILE": str(workbook_path)}):
                self.module["_enhance_and_filter"](category_c, lambda _message: None)
                self.module["_enhance_and_filter"](category_b, lambda _message: None)

            workbook = openpyxl.load_workbook(workbook_path, data_only=True)
            self.assertEqual(
                workbook.sheetnames,
                ["Sheet1", category_b["data_sheet"], category_c["data_sheet"]],
            )
            workbook.close()

    def test_pivot_cache_uses_an_activated_contiguous_source_sheet(self):
        class FakeField:
            Orientation = None
            Position = None

        class FakePivotTable:
            def PivotFields(self, _name):
                return FakeField()

            def AddDataField(self, *_args):
                return None

        class FakePivotCache:
            def CreatePivotTable(self, **_kwargs):
                return FakePivotTable()

        class FakePivotCaches:
            def __init__(self):
                self.create_kwargs = None

            def Create(self, **kwargs):
                self.create_kwargs = kwargs
                return FakePivotCache()

        class FakeCells:
            def __call__(self, *_args):
                return object()

        class FakeSheet:
            def __init__(self, name):
                self.Name = name
                self.Cells = FakeCells()
                self.activated = False

            def Activate(self):
                self.activated = True

        class FakeSheets:
            def __init__(self, sheets):
                self._sheets = sheets
                self.add_after = None

            def __iter__(self):
                return iter(self._sheets)

            def __call__(self, name):
                return next(sheet for sheet in self._sheets if sheet.Name == name)

            def Add(self, After):
                self.add_after = After
                sheet = FakeSheet("new pivot")
                self._sheets.append(sheet)
                return sheet

        source_sheet = FakeSheet("DEMO_CATEGORY_A_DATA")
        sheets = FakeSheets([FakeSheet("Sheet1"), source_sheet])
        pivot_caches = FakePivotCaches()
        workbook = types.SimpleNamespace(
            Name="demo_purchase_records.xlsx",
            Sheets=sheets,
            PivotCaches=lambda: pivot_caches,
        )

        self.module["_com_create_pivot"](
            workbook,
            total_cols=5,
            total_rows=1,
            source_name=source_sheet.Name,
            target_sheet="DEMO_CATEGORY_A_PIVOT",
            pivot_table_name="DEMO_CATEGORY_A_PIVOT",
            row_fields=["区域"],
            value_field="订单净值",
            value_name="求和项:订单净值",
        )

        self.assertIs(sheets.add_after, source_sheet)
        self.assertTrue(source_sheet.activated)
        self.assertEqual(
            pivot_caches.create_kwargs["SourceData"],
            "'[demo_purchase_records.xlsx]DEMO_CATEGORY_A_DATA'!R1C1:R2C5",
        )

    def test_configuration_template_is_parseable_and_uses_placeholder_urls(self):
        template_path = PROJECT_DIR / "config.example.json"
        template = json.loads(template_path.read_text(encoding="utf-8"))

        self.assertEqual(set(template["web"]), {"home_url", "search_url"})
        for value in template["web"].values():
            self.assertIn(".example", value)

    def test_missing_runtime_urls_fail_before_any_browser_workflow(self):
        with self.assertRaises(ConfigurationError):
            load_web_urls(config_path=PROJECT_DIR / "missing-config.json", environ={})


if __name__ == "__main__":
    unittest.main()
